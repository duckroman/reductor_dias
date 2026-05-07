import openpyxl
import random
import numpy as np
import os

filepath = 'cumplimiento_completo_v1.xlsx'

if not os.path.exists(filepath):
    print(f"Error: {filepath} no existe.")
    exit(1)

print(f"Cargando {filepath}...")
wb = openpyxl.load_workbook(filepath)

# 90% reach 100 on Day 45, 10% on Day 56
target_days_pool = [45] * 90 + [56] * 10

for sheet_name in wb.sheetnames:
    print(f"Procesando hoja: {sheet_name}")
    ws = wb[sheet_name]
    
    # Encontrar la fila de encabezados para confirmar columnas
    # Según la inspección, los datos empiezan en la fila 3 (index 3)
    # y las columnas de días empiezan en la 5 (E)
    
    max_row = ws.max_row
    # El archivo tiene 300 distritos por hoja aproximadamente
    
    for row_idx in range(3, max_row + 1):
        # Asignar un día objetivo aleatorio basado en la distribución 90/10
        target_day = random.choice(target_days_pool)
        
        # Parámetros para la curva logística
        # Queremos que en target_day llegue a ~1.0
        # C(t) = L / (1 + exp(-k * (t - t0)))
        # t0 será alrededor de 20-25
        t0 = random.uniform(15, 25)
        k = random.uniform(0.15, 0.25)
        
        last_val = 0.0
        for day in range(1, 57):
            col_idx = 5 + (day - 1)
            
            # Si el día es mayor o igual al objetivo, es 100%
            if day >= target_day:
                val = 1.0
            else:
                # Curva logística base
                val = 1.0 / (1.0 + np.exp(-k * (day - t0)))
                
                # Ajuste para que empiece cerca de 0 y termine cerca de 1 en target_day
                # Pero la logística natural ya hace eso.
                # Solo aseguramos que sea creciente y tenga un poco de ruido
                val += random.uniform(0, 0.01)
                val = max(last_val, min(0.99, val)) # No llegar a 1 antes de tiempo
            
            last_val = val
            ws.cell(row=row_idx, column=col_idx, value=val)
            
    print(f"Hoja {sheet_name} completada.")

output_filepath = 'cumplimiento_completo_v1_poblado.xlsx'
print(f"Guardando cambios en {output_filepath}...")
wb.save(output_filepath)
print("¡Proceso finalizado con éxito!")
