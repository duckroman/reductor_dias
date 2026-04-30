"""
Generador de datos sintéticos para cumplimiento de visitas - INE
Genera curvas logísticas (S-curve) con parámetros aleatorios para 300 distritos × 50 días.
"""

import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# Semilla para reproducibilidad
np.random.seed(42)

NUM_DISTRITOS = 300
NUM_DIAS = 50

def logistic_curve(t, L, k, x0):
    """Curva logística: L / (1 + exp(-k*(t - x0)))"""
    return L / (1 + np.exp(-k * (t - x0)))

def generar_datos():
    """Genera la matriz de datos sintéticos con curvas logísticas."""
    dias = np.arange(1, NUM_DIAS + 1)
    datos = np.zeros((NUM_DISTRITOS, NUM_DIAS))
    
    for i in range(NUM_DISTRITOS):
        # Parámetros aleatorios por distrito
        L = np.random.uniform(0.75, 1.0)       # Asíntota máxima (75%-100%)
        k = np.random.uniform(0.05, 0.20)       # Velocidad de crecimiento
        x0 = np.random.uniform(15, 35)           # Día de inflexión (punto medio)
        
        # Generar la curva base
        curva = logistic_curve(dias, L, k, x0)
        
        # Agregar ruido gaussiano (±1-3%)
        noise_level = np.random.uniform(0.01, 0.03)
        ruido = np.random.normal(0, noise_level, NUM_DIAS)
        curva_con_ruido = curva + ruido
        
        # Forzar monotonía creciente
        for j in range(1, NUM_DIAS):
            if curva_con_ruido[j] < curva_con_ruido[j - 1]:
                curva_con_ruido[j] = curva_con_ruido[j - 1] + np.random.uniform(0.001, 0.005)
        
        # Limitar entre 0 y 1
        curva_con_ruido = np.clip(curva_con_ruido, 0.0, 1.0)
        
        datos[i] = curva_con_ruido
    
    return datos

def escribir_excel(datos, filepath):
    """Escribe los datos en el archivo Excel con formato profesional."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    # --- Estilos ---
    header_font = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    
    col_header_font = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
    col_header_fill = PatternFill(start_color="3A6BC5", end_color="3A6BC5", fill_type="solid")
    
    data_font = Font(name='Calibri', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # --- Fila 1: Título ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_DIAS + 1)
    title_cell = ws.cell(row=1, column=1, value="Porcentaje de cumplimiento en visitas a ciudadanos")
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30
    
    # --- Fila 2: Encabezados de columnas ---
    ws.cell(row=2, column=1, value="Distrito").font = col_header_font
    ws.cell(row=2, column=1).fill = col_header_fill
    ws.cell(row=2, column=1).alignment = center_align
    ws.cell(row=2, column=1).border = thin_border
    ws.column_dimensions['A'].width = 12
    
    for d in range(1, NUM_DIAS + 1):
        col = d + 1
        cell = ws.cell(row=2, column=col, value=f"Día {d}")
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 9
    
    ws.row_dimensions[2].height = 22
    
    # --- Filas 3–302: Datos ---
    # Colores alternados para filas
    fill_even = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for i in range(NUM_DISTRITOS):
        row = i + 3
        row_fill = fill_even if i % 2 == 0 else fill_odd
        
        # Columna Distrito
        cell = ws.cell(row=row, column=1, value=i + 1)
        cell.font = Font(name='Calibri', bold=True, size=10)
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = thin_border
        
        # Columnas de días
        for d in range(NUM_DIAS):
            col = d + 2
            cell = ws.cell(row=row, column=col, value=round(datos[i][d], 4))
            cell.font = data_font
            cell.alignment = center_align
            cell.number_format = '0.00%'
            cell.fill = row_fill
            cell.border = thin_border
    
    # --- Congelar paneles ---
    ws.freeze_panes = 'B3'
    
    # --- Guardar ---
    wb.save(filepath)
    print(f"[OK] Archivo guardado: {filepath}")
    print(f"   Distritos: {NUM_DISTRITOS}")
    print(f"   Días: {NUM_DIAS}")
    print(f"   Total celdas de datos: {NUM_DISTRITOS * NUM_DIAS:,}")

def validar_datos(datos):
    """Valida las propiedades de los datos generados."""
    print("\n[STATS] Validacion de datos:")
    print(f"   Rango de valores: [{datos.min():.4f}, {datos.max():.4f}]")
    print(f"   Media global: {datos.mean():.4f}")
    print(f"   Mediana global: {np.median(datos):.4f}")
    
    # Verificar monotonía creciente
    mono_ok = 0
    for i in range(NUM_DISTRITOS):
        diffs = np.diff(datos[i])
        if np.all(diffs >= 0):
            mono_ok += 1
    print(f"   Distritos monótonos crecientes: {mono_ok}/{NUM_DISTRITOS}")
    
    # Estadísticas por columnas representativas
    for day_idx in [0, 9, 24, 39, 49]:
        col = datos[:, day_idx]
        print(f"   Día {day_idx+1}: media={col.mean():.3f}, min={col.min():.3f}, max={col.max():.3f}")

if __name__ == "__main__":
    print("[INFO] Generando datos sinteticos para cumplimiento de visitas...")
    datos = generar_datos()
    validar_datos(datos)
    
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cumplimiento_visitas.xlsx")
    try:
        escribir_excel(datos, filepath)
    except PermissionError:
        print("[WARN] No se pudo escribir en el archivo original (puede estar abierto en Excel).")
        backup = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cumplimiento_visitas_nuevo.xlsx")
        escribir_excel(datos, backup)
        print(f"[INFO] Se guardo como: {backup}")
        print("[INFO] Cierra Excel y renombra el archivo manualmente, o ejecuta de nuevo.")
    print("\n[OK] Proceso completado!")
